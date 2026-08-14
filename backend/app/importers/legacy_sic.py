from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CANONICAL_HEADERS = {
    "codigo": "legacy_code",
    "código": "legacy_code",
    "produto": "description",
    "descricao": "description",
    "descrição": "description",
    "quantidade": "current_stock",
    "qtd": "current_stock",
    "estoque": "current_stock",
    "valor": "sale_price",
    "preco": "sale_price",
    "preço": "sale_price",
}


@dataclass(frozen=True)
class NormalizedPartPreview:
    row_number: int
    legacy_code: str
    description: str | None
    current_stock: Decimal
    sale_price: Decimal
    raw: dict[str, Any]


def normalize_header(header: Any) -> str:
    return str(header or "").strip().lower()


def normalize_code(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int | float):
        return Decimal(str(value))
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


class LegacySICImporter:
    """Read-only analyzer for the SIC product export.

    The importer deliberately stops at analysis/preview. A future confirmed import
    can reuse the normalized rows and persist them into parts + stock_movements.
    """

    def __init__(self, path: str | Path):
        self.path = Path(path)

    def analyze(self, preview_limit: int = 10) -> dict[str, Any]:
        if not self.path.exists():
            return {"workbook": str(self.path), "exists": False, "sheets": []}

        workbook = load_workbook(self.path, read_only=True, data_only=True)
        sheets: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            sheets.append(self._analyze_sheet(worksheet, preview_limit=preview_limit))
        return {"workbook": str(self.path), "exists": True, "sheets": sheets}

    def _analyze_sheet(self, worksheet: Any, *, preview_limit: int) -> dict[str, Any]:
        header_row, headers = self._detect_header(worksheet)
        if not headers:
            return {
                "name": worksheet.title,
                "empty": True,
                "max_row": worksheet.max_row,
                "max_col": worksheet.max_column,
            }

        rows = self._read_data_rows(worksheet, header_row, len(headers))
        mapped_headers = self._map_headers(headers)
        previews, row_issues = self._build_preview(rows, headers, mapped_headers, preview_limit)
        column_stats = self._column_stats(rows, headers)
        code_idx = mapped_headers.get("legacy_code")
        description_idx = mapped_headers.get("description")
        price_idx = mapped_headers.get("sale_price")
        stock_idx = mapped_headers.get("current_stock")

        inconsistencies: dict[str, Any] = {
            "rows_with_any_issue": len(row_issues),
            "issue_examples": row_issues[:25],
        }
        if code_idx is not None:
            codes = [normalize_code(row[code_idx]) for _, row in rows if normalize_code(row[code_idx])]
            duplicate_codes = [code for code, count in Counter(codes).items() if count > 1]
            inconsistencies["duplicated_codes"] = len(duplicate_codes)
            inconsistencies["duplicated_code_examples"] = duplicate_codes[:25]
            inconsistencies["empty_codes"] = len(rows) - len(codes)
        if description_idx is not None:
            empty_description_rows = [
                row_number
                for row_number, row in rows
                if row[description_idx] in (None, "") or not str(row[description_idx]).strip()
            ]
            inconsistencies["empty_descriptions"] = len(empty_description_rows)
            inconsistencies["empty_description_rows"] = empty_description_rows[:25]
        if price_idx is not None:
            inconsistencies["invalid_prices"] = self._numeric_issue_count(rows, price_idx, negative_allowed=False)
        if stock_idx is not None:
            inconsistencies["invalid_quantities"] = self._numeric_issue_count(rows, stock_idx, negative_allowed=True)
            negative_rows = [
                row_number
                for row_number, row in rows
                if (value := parse_decimal(row[stock_idx])) is not None and value < 0
            ]
            inconsistencies["negative_quantities"] = len(negative_rows)
            inconsistencies["negative_quantity_rows"] = negative_rows[:25]

        return {
            "name": worksheet.title,
            "empty": False,
            "max_row": worksheet.max_row,
            "max_col": worksheet.max_column,
            "header_row": header_row,
            "headers": headers,
            "mapped_headers": {key: headers[index] for key, index in mapped_headers.items()},
            "records": len(rows),
            "column_stats": column_stats,
            "inconsistencies": inconsistencies,
            "preview": previews,
        }

    def _detect_header(self, worksheet: Any) -> tuple[int, list[str]]:
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = list(row or [])
            if any(value not in (None, "") for value in values):
                return row_number, [
                    str(value).strip() if value is not None else f"col_{index + 1}"
                    for index, value in enumerate(values)
                ]
        return 0, []

    def _read_data_rows(
        self,
        worksheet: Any,
        header_row: int,
        width: int,
    ) -> list[tuple[int, list[Any]]]:
        data_rows: list[tuple[int, list[Any]]] = []
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            values = list(row or [])[:width]
            values = values + [None] * (width - len(values))
            if any(value not in (None, "") for value in values):
                data_rows.append((row_number, values))
        return data_rows

    def _map_headers(self, headers: list[str]) -> dict[str, int]:
        mapped: dict[str, int] = {}
        for index, header in enumerate(headers):
            canonical = CANONICAL_HEADERS.get(normalize_header(header))
            if canonical and canonical not in mapped:
                mapped[canonical] = index
        return mapped

    def _column_stats(self, rows: list[tuple[int, list[Any]]], headers: list[str]) -> list[dict[str, Any]]:
        stats: list[dict[str, Any]] = []
        for column_index, header in enumerate(headers):
            values = [row[column_index] for _, row in rows]
            non_empty = [value for value in values if value not in (None, "")]
            stats.append(
                {
                    "header": header,
                    "non_empty": len(non_empty),
                    "empty": len(values) - len(non_empty),
                    "types": dict(Counter(type(value).__name__ for value in non_empty)),
                    "sample": [str(value) for value in non_empty[:5]],
                }
            )
        return stats

    def _build_preview(
        self,
        rows: list[tuple[int, list[Any]]],
        headers: list[str],
        mapped_headers: dict[str, int],
        preview_limit: int,
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        preview: list[dict[str, Any]] = []
        issues: list[dict[str, Any]] = []
        for row_number, row in rows:
            raw = {headers[index]: row[index] for index in range(len(headers))}
            normalized = {
                "row_number": row_number,
                "legacy_code": normalize_code(row[mapped_headers["legacy_code"]])
                if "legacy_code" in mapped_headers
                else "",
                "description": str(row[mapped_headers["description"]]).strip()
                if "description" in mapped_headers and row[mapped_headers["description"]] not in (None, "")
                else None,
                "current_stock": parse_decimal(row[mapped_headers["current_stock"]])
                if "current_stock" in mapped_headers
                else None,
                "sale_price": parse_decimal(row[mapped_headers["sale_price"]])
                if "sale_price" in mapped_headers
                else None,
                "raw": raw,
            }
            row_issue_codes = self._row_issues(normalized)
            if row_issue_codes:
                issues.append({"row_number": row_number, "issues": row_issue_codes, "raw": raw})
            if len(preview) < preview_limit:
                preview.append(self._serialize_preview(normalized))
        return preview, issues

    def _row_issues(self, normalized: dict[str, Any]) -> list[str]:
        issues: list[str] = []
        if not normalized["legacy_code"]:
            issues.append("CODIGO_VAZIO")
        if not normalized["description"]:
            issues.append("DESCRICAO_VAZIA")
        if normalized["sale_price"] is None:
            issues.append("PRECO_INVALIDO_OU_VAZIO")
        if normalized["current_stock"] is None:
            issues.append("QUANTIDADE_INVALIDA_OU_VAZIA")
        elif normalized["current_stock"] < 0:
            issues.append("QUANTIDADE_NEGATIVA")
        return issues

    def _numeric_issue_count(
        self,
        rows: list[tuple[int, list[Any]]],
        column_index: int,
        *,
        negative_allowed: bool,
    ) -> dict[str, Any]:
        invalid_rows: list[int] = []
        negative_rows: list[int] = []
        for row_number, row in rows:
            value = parse_decimal(row[column_index])
            if row[column_index] not in (None, "") and value is None:
                invalid_rows.append(row_number)
            if value is not None and value < 0 and not negative_allowed:
                negative_rows.append(row_number)
        return {
            "invalid": len(invalid_rows),
            "invalid_rows": invalid_rows[:25],
            "negative": len(negative_rows),
            "negative_rows": negative_rows[:25],
        }

    def _serialize_preview(self, normalized: dict[str, Any]) -> dict[str, Any]:
        return {
            **normalized,
            "current_stock": str(normalized["current_stock"])
            if normalized["current_stock"] is not None
            else None,
            "sale_price": str(normalized["sale_price"])
            if normalized["sale_price"] is not None
            else None,
        }
